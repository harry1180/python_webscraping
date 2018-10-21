import datetime
import urllib3
import requests
import os
from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd
import logging
from openpyxl import Workbook,load_workbook


'''*********************************************Parameters**********************************************************************'''
maindirectory = os.path.realpath('Google')
searchdata = "searchcriteriadata.xlsx"
searchCriteriaInput = os.path.join(maindirectory,searchdata)
searchResults = 'searchResults_1510.xlsx'
resultsOutput = os.path.join(maindirectory,searchResults)
webdriverpath = os.path.join('chromedriver.exe')
browserName = "Chrome"
googlepage = "https://www.google.com"
gt_logo = "gt_logo.png"


urls = []
headlines = []
articles = []

'''************************************Initialise Log file************************************************************************'''
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Create a file handler

handler = logging.FileHandler('log.ini')
handler.setLevel(logging.INFO)

# Create a logging format
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

# Add the handlers to the logger
logger.addHandler(handler)





class googleSearch:

    def __init__(self):
        logger.info("______Initializing Window________")


    '''***************************Step 1 : Open Broswser Window******************************'''
    def openBrowser(self):
        logger.info("______Initializing Web Browser________")
        options = webdriver.ChromeOptions()
        options.binary_location = webdriverpath
        self.browser = webdriver.Chrome(executable_path=webdriverpath)
        return self.browser

    ''' ***************************Step 2 : Navigate to Google Search******************************'''

    def  openGoogle(self,googlepage):
        logger.info(("______Opening Google Search_________"))
        searchpage = self.browser.get(googlepage)
        return searchpage

    ''' ***************************Step 3 : Read the Search criteria******************************'''
    def loadSearchCriteria(self,limiteddata):
        logger.info("_______Reading Search Criteria________")
        #searchdataframe = pd.read_excel(searchCriteriaInput)
        global searchdataframe,jurisdiction,searchcriteria,results,searchByCountry

        searchdataframe = pd.read_excel(searchdata, sheet_name='Sheet1')

        jurisdiction = searchdataframe["Jurisdiction"].unique().tolist()
        searchcriteria = searchdataframe["Search Criteria"].unique().tolist()
        logger.info("Number of Jurisdictions : " + str(len(jurisdiction)))
        logger.info("Number of items to search : " + str(len(searchcriteria)))
        # print(len(searchcriteria))
        rowLen = "Number of Jurisdictions : " + str(len(jurisdiction))
        colLen = "Number of items to search : " + str(len(searchcriteria))

        return jurisdiction,searchcriteria,searchdataframe

    ''' ***************************Step 4 : Initiate the Search ******************************'''
    def initiateSearch(self, searchpage,country,criteria):
        logger.info("______Initiate Google Search________")

        #Locate the Search Box
        searchBox = self.browser.find_element_by_xpath("//*[@id='lst-ib']")

        #Type the search criteria in the search Box
        searchBox.send_keys(criteria)
        searchBox.submit()
        self.browser.find_element_by_link_text("News").click()
        soup = BeautifulSoup(self.browser.page_source, 'html.parser')
        # print(soup.prettify())
        # self.browser.implicitly_wait(5)

        #Locate the search results
        news_list = soup.find_all('div', class_="gG0TJc")
        #print(news_list)

        l = []
        d = {}
        dfs = dict()
        articles=[]
        urls=[]
        headlines=[]
        '''For each item in the searchPage capture the link for the newitem,newsheadline and description text.
         captures only the top 10 or first page of the results'''
        for item in news_list:
            # d={}

            d["Jurisdiction"] = country
            #print("For item in news list picked country",country)
            d["Criteria"] = criteria

            ''' if (d["Criteria"].find(jurisdiction) != -1):
                d.update({"Country": jurisdiction})'''

            news_link = item.find('a', href=True)
            urls.append(news_link.attrs['href'])
            d["News Links"] = urls
            
            news_headline = news_link.text
            headlines.append(news_headline)
            d["News Headlines"] = headlines


            news_article = item.find('div', class_="st")
            articles.append((news_article).text)
            d["News Articles"] = articles



        self.browser.implicitly_wait(10)

        searchByCountry = str(d.get('Jurisdiction'))
        print(searchByCountry,'this is search by country')
        dframe = pd.DataFrame.from_dict(d)
        print(d,'this should be dictionary')

        # dframe.append(result)
        # dframe.to_excel('output.xlsx')
        
        
        self.append_df_to_excel(dframe,searchByCountry)
        self.closeBrowser()

        
        return searchByCountry,dframe

    ''' ***************************Step 5 : write the search results to excel sheet.the format expected to be to write the results 
            per jurisdiction.Please verify the sample output xls file ******************************'''
    
    def abc():
        data1=pd.read_json('amrutha_op.json')
        print(data1.describe)




    def closeBrowser(self):
        self.browser.quit()


    def append_df_to_excel(self,dframe,searchByCountry,truncate_sheet = False,startrow = None,**to_excel_kwargs):
        writer = pd.ExcelWriter(searchResults,engine = 'openpyxl')
        sheets=[]
        print (searchResults,'this should be file path')
        '''Check if the file already exists'''
        if(os.path.exists(searchResults)):
            print("File exists")
            writer.book = load_workbook(searchResults)
            print(writer.book,'this is writer.book')
            print(writer.book.sheetnames)
            '''Issue callinng the worksheet as it is retrieving the worksheet object as <worksheet_UK>.So converted the worksheet Object to string.
             /Verifying if te worksheet name is same as the SearchByCountry'''
            for ws in writer.book.sheetnames:
                print(ws,'this is our interest')
                sheetname = str(writer.book[ws])
                import re
                #matches=re.findall(r'\"(.+?\D)\"',sheetname)
                matches=re.findall(r'\".*\"',sheetname)
                aa=','.join(matches)
                b=aa.strip('"')
                striped=b.rstrip('0123456789')

                # try:
                wsname = striped
                # except:
                #     pass
                # else:
                #     wsname = aa.split(' ')[0].strip('"')
                print("1 :",sheetname)
                #wsname = sheetname.split(" ")[1].strip(("', >"))
                print("2 :",wsname)
                # worksheet = (wsname.strip('"'))
                worksheet = wsname
                print(worksheet,'this is worksheet')
               # print("3 :",worksheet)
                sheets.append(worksheet)
                '''If Worksheet already exists then get the maxrow of the worksheet and write to excel starting at maxrow'''
            print(sheets,'this sheet elements together')
            for wname in sheets:
                print(wname,'should be worksheet elements')
                # if (searchByCountry == wname):
                if (searchByCountry in wname):
                    searchByCountry=wname
                    startrow = writer.book[worksheet].max_row
                    print(startrow,'this is startrow')
                    print(worksheet,'this is worksheet before getting to excel')
                    dframe.to_excel(writer, worksheet, startrow, index =False)
                    writer.save()

                    

                else:
                    '''If Worksheet does not exist then get create a new sheet and write to excel starting at startrow'''
                    print('we are in else')
                    startrow = writer.book[worksheet].max_row
                    print(startrow,'this is startrow in else')
                    dframe.to_excel(writer,searchByCountry,startrow, index =False)
                    # dframe.append(writer,searchByCountry)

        else:
            #print("File does not exists")
            writer = pd.ExcelWriter(searchResults, engine='openpyxl')
            dframe.to_excel(writer,searchByCountry,startrow, index =False)
            # dframe.to_excel(writer,searchByCountry)
        writer.save()
        self.closeBrowser()




    def closeBrowser(self):
        self.browser.quit()


def main():

    starttime = datetime.datetime.now()
    print(starttime)


    google = googleSearch()
    jurisdiction,searchcriteria,searchdataframe=google.loadSearchCriteria(searchdata)
    for criteria in searchcriteria:
        google.openBrowser()
        googleresponse = google.openGoogle(googlepage)
        print(len(jurisdiction),'this is count of jurisdiction')
        for country in jurisdiction:
            if(criteria.find(country)!=-1) :
                urls = []
                headlines = []
                articles = []
                print("Picked up country",country)
                print(criteria,'thisis creteria')
                #results,searchByCountry = google.initiateSearch(googleresponse, country, criteria)
                Jurisdiction,searchByCountry = google.initiateSearch(googleresponse, country, criteria)
                print(Jurisdiction,'this is results response')
                print(searchByCountry,'this is searchByCountry response')
                google.closeBrowser()



    endtime = datetime.datetime.now()
    print(endtime)




if __name__ == '__main__':
    main()

